import gc
import json
import re
from scripts_public import _setup_django
from unicodedata import decimal
from openpyxl import load_workbook


def validate_red_upload_excel_row(index, row_data):
    rep_row_data = {
        'line_index': index,
        'origin_code': '',  # 原始单号
        'order_code': '',  # 解析单号
        'check_success': True,
    }

    origin_code = row_data.get('单号')
    rep_row_data['origin_code'] = origin_code
    # 去掉空格
    if origin_code:
        order_code = origin_code.replace(' ', '')
        rep_row_data['order_code'] = order_code
        if not order_code:
            rep_row_data['check_success'] = False
    else:
        rep_row_data['check_success'] = False
    return rep_row_data


class Read_RED_Upload_Excel(object):
    """
    读取验证excel文件数据
    """

    def __init__(self, file_path, col_field=None, header_index=1):
        self.file_path = file_path
        self.header_index = header_index
        self.err_msg = ''
        self.col_field = col_field
        self.row_func = validate_red_upload_excel_row

    def validate_col(self, col_keys):
        """
        表头必填列验证
        :param col_keys:
        :return:
        """
        if self.col_field:
            col_field = set(self.col_field)
            fields_set = col_field & set(col_keys)
            return col_field - fields_set

    def validate_row(self, index, row_data):
        """
        行数据验证
        :param index:
        :param row_data:
        :return:
        """
        if self.row_func:
            row_data = self.row_func(index, row_data)
        return row_data

    def read_data(self):
        """
        :return:
        """
        index = self.header_index - 1
        return self.read_xlsx(index)

    def read_xlsx(self, header_index):
        """
        @summary: 从Excel文件中读取数据后缀(.xlsx)
        @return: 返回数据格式 [{},{},]
        """
        workbook = load_workbook(self.file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        data = []
        max_row = worksheet.max_row
        if max_row > 1:
            rows = worksheet.rows
            key_word = []  # 首行信息
            for index, row in enumerate(rows):
                if index == header_index:
                    for col in row:
                        val = str(col.value or '').strip().replace('\n', '')
                        key_word.append(val)
                    col_field = self.validate_col(key_word)
                    if col_field:
                        self.err_msg = f'上传excel模板缺少必填列{col_field}， 请使用正确的模板导入！'
                        break
                elif index > header_index:
                    data_dic = {}
                    for i, col in enumerate(row):
                        data_dic[key_word[i]] = col.value
                    if not any(data_dic.values()):
                        continue
                    row_data = self.validate_row(index, data_dic)
                    if row_data:
                        data.append(row_data)
        del workbook, worksheet
        gc.collect()
        return self.err_msg, data


if __name__ == '__main__':
    test_file_path = r'D:\workProjectCode\ReserveDownloadSignleBackend\MEDIA_ROOT\RED_Upload_Excel\6623_1713595488_4544_mobj.xlsx'
    col_field_test = ['单号']

    t = Read_RED_Upload_Excel(test_file_path, col_field_test)
    err, data = t.read_data()
    print('err:', err)
    print('data:', data)